from decimal import Decimal, InvalidOperation
import re, time, unicodedata
from datetime import datetime, date, timedelta

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Empresa, LoginLog, SatRegistro


# ------------------ Utils básicos ------------------

def _slugify_field(name: str) -> str:
    if name is None:
        name = ""
    name = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    name = re.sub(r"[^\w]+", "_", name.strip().lower()).strip("_")
    if not name:
        name = "campo"
    if re.match(r"^\d", name):
        name = f"col_{name}"
    if name in {"class","def","return","yield","from","import","global",
                "lambda","with","pass","raise","id","pk","model","objects"}:
        name = f"{name}_field"
    return name

def _slug(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^\w]+", "_", s.strip().lower()).strip("_") or "campo"

def _first_key(d: dict, candidates):
    for k in candidates:
        if k in d:
            return k
    low_map = {k.lower(): k for k in d.keys()}
    for c in candidates:
        for klow, korig in low_map.items():
            if c in klow:
                return korig
    return None

def _parse_decimal(v):
    if v is None:
        return Decimal("0")
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s and s.count(",") == 1:
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


# ------------------ Normalização/extrações ------------------

def _is_cancelado(txt: str) -> bool:
    return bool(txt) and "cancel" in str(txt).strip().lower()

def _extr_status_sat(d: dict) -> str:
    k = _first_key(d, ["situacao", "status", "status_nfce", "situacao_nfce"])
    return str(d.get(k, "")).strip().lower() if k else ""

def _extr_valor_total_sat(d: dict) -> Decimal:
    k = _first_key(d, ["valornfe","valor_total","valortotal","valor_total_nfe","valor_nfce"])
    return _parse_decimal(d.get(k))

def _extr_id_sat(d: dict) -> str:
    for k in ["chaveacesso","chave_de_acesso","chave","accesskey"]:
        if d.get(k):
            return str(d[k])
    num = d.get(_first_key(d, ["numerodocumento","numero","numnfe","numero_nfce"])) or ""
    serie = d.get(_first_key(d, ["serie","serienfe","serie_nfce"])) or ""
    return (f"Nº {num} • Série {serie}").strip(" •") or "(sem id)"

def _numero_documento_sat(d: dict) -> str:
    k = _first_key(d, ["cu_numerodocumento","numerodocumento","numero_documento",
                       "documento","n_documento","numnfe","numero_nfce","numero"])
    return (str(d.get(k)) if k and d.get(k) is not None else "").strip()

def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "").strip()

def _norm_doc(s: str) -> tuple[str, str]:
    raw = (s or "").strip().lower()
    digs = _digits_only(s)
    return raw, digs

def _norm_pair(doc: str, serie: str) -> tuple[str, str]:
    raw_doc = (doc or "").strip().lower()
    raw_ser = (serie or "").strip().lower()
    dig_doc = _digits_only(doc)
    dig_ser = _digits_only(serie)
    return f"{raw_doc}|{raw_ser}", f"{dig_doc}|{dig_ser}"

def _norm_txt(s: str) -> str:
    t = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return t.lower().replace(" ", "").replace("-", "").replace(".", "")

def _is_nfe_questor(d: dict) -> bool:
    k_esp = _first_key(d, ["especie","especie_documento","tipo","tipo_documento"])
    if k_esp:
        v = _norm_txt(d.get(k_esp, ""))
        if "nfce" in v or "nfc" in v or v == "65":
            return False
        if "nfe" in v or "nfeletronica" in v or v == "55":
            return True
    k_mod = _first_key(d, ["modelodocumento","modelo_documento","modelo","mod",
                           "modelo_nf","modelo_nfe","modelo_nfce"])
    if k_mod:
        v = _norm_txt(d.get(k_mod, ""))
        if "65" in v or "nfce" in v or "nfc" in v:
            return False
        if "55" in v or "nfe" in v:
            return True
    k_ch = _first_key(d, ["chaveacesso","chave_de_acesso","chave","accesskey"])
    if k_ch:
        digs = _digits_only(d.get(k_ch, ""))
        if len(digs) >= 22:
            modelo = digs[20:22]
            if modelo == "55":
                return True
            if modelo == "65":
                return False
    return False

def _is_autorizado(txt: str) -> bool:
    t = (txt or "").strip().lower()
    if not t or _is_cancelado(t):
        return False
    tokens = ["autoriz","aprov","normal","regular","emitid"]
    return any(tok in t for tok in tokens)

def _status_legivel(d: dict) -> str:
    s = _extr_status_sat(d)
    if _is_cancelado(s):
        return "CANCELADA"
    if _is_autorizado(s):
        return "AUTORIZADA"
    return (s or "DESCONHECIDO").strip().upper()


# ------------------ Datas e competência ------------------

_DT_FMTS = (
    "%d/%m/%y","%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%d.%m.%Y",
    "%Y-%m-%d %H:%M","%Y-%m-%d %H:%M:%S",
    "%d/%m/%Y %H:%M","%d/%m/%Y %H:%M:%S",
    "%d/%m/%y %H:%M","%d/%m/%y %H:%M:%S"
)

def _parse_date_any(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    s_norm = s.replace("T", " ").replace("Z", "")
    if "." in s_norm:
        s_norm = s_norm.split(".")[0]
    for fmt in _DT_FMTS:
        try:
            return datetime.strptime(s_norm, fmt).date()
        except Exception:
            pass
    if s.isdigit():
        try:
            n = int(s)
            if 20000 <= n <= 80000:
                return date(1899, 12, 30) + timedelta(days=n)
        except Exception:
            pass
    return None

def _extr_data_emissao_dict(d: dict) -> date | None:
    k = _first_key(d, [
        "dataemissao","data_emissao","dt_emissao","emissao","emissao_data",
        "dataemissaonf","data_emissao_nf","data","dtemissao",
        "data_entrada_saida","data_entrada","data_saida","dt_entrada","dt_saida",
        "dataentradasaida"
    ])
    return _parse_date_any(d.get(k)) if k else None

def _competencia_from_date(d: date | None) -> date | None:
    return date(d.year, d.month, 1) if d else None

def _parse_competencia_str(s: str | None) -> date | None:
    if not s:
        return None
    t = s.strip()
    try:
        if "-" in t:  # YYYY-MM
            y, m = t.split("-", 1)
            return date(int(y), int(m), 1)
        if "/" in t:  # MM/YYYY
            m, y = t.split("/", 1)
            return date(int(y), int(m), 1)
    except Exception:
        return None
    return None


# ------------------ Auth e páginas simples ------------------

def login_view(request):
    if request.method == 'POST':
        email = (request.POST.get('email') or '').strip().lower()
        password = request.POST.get('password') or ''
        user = None
        try:
            user = authenticate(request, email=email, password=password)
        except TypeError:
            user = None
        if user is None and email and password:
            User = get_user_model()
            u = User.objects.filter(email__iexact=email).order_by('id').first()
            if u:
                user = authenticate(request, username=u.get_username(), password=password)
        try:
            LoginLog.objects.create(user=user if user else None, email=email, success=bool(user))
        except Exception:
            pass
        if user is not None:
            if not user.is_active:
                messages.error(request, 'Usuário inativo.')
                return render(request, 'login.html', {'hide_navbar': True})
            login(request, user)
            return redirect(request.GET.get('next') or 'painel_inicial')
        messages.error(request, 'E-mail ou senha inválidos.')
    return render(request, 'login.html', {'hide_navbar': True})

@login_required
def logout_view(request):
    if request.method == "POST":
        logout(request)
        messages.success(request, "Você saiu do sistema.")
        return redirect("login")
    return redirect("painel_inicial")

@login_required
def painel_inicial(request):
    empresas = Empresa.objects.all().order_by("nome")
    return render(request, "painel_inicial.html", {"empresas": empresas})


# ------------------ IMPORTAÇÃO SAT (usa competência) ------------------

@login_required
@require_POST
@transaction.atomic
def sat_importar(request):
    """
    Importa XLSX do SAT com UPSERT pela chave:
    (empresa, competencia, sheet, row).
    - Se mudar a competência (mês), será INSERT.
    - Se for a mesma competência, será UPDATE.
    Opcional: campo 'competencia' vindo do form (YYYY-MM ou MM/YYYY)
    caso a planilha não tenha data de emissão por linha.
    """
    empresa_id = request.POST.get("empresa_id")
    arquivo    = request.FILES.get("arquivo")
    comp_param = _parse_competencia_str(request.POST.get("competencia"))

    if not empresa_id:
        messages.error(request, "Selecione a empresa.")
        return redirect("painel_inicial")
    if not arquivo or not arquivo.name.lower().endswith(".xlsx"):
        messages.error(request, "Envie um arquivo .xlsx para importar.")
        return redirect("painel_inicial")

    empresa = get_object_or_404(Empresa, pk=empresa_id)

    try:
        wb = load_workbook(filename=arquivo, data_only=True, read_only=True)
    except Exception as e:
        messages.error(request, f"Falha ao abrir o Excel: {e}")
        return redirect("painel_inicial")

    inicio = time.time()
    BATCH = 500
    criados = atualizados = ignorados_vazios = 0
    to_create: list[SatRegistro] = []
    to_update: list[SatRegistro] = []

    # mapa existentes: (empresa, competencia, sheet, row)
    exist_map: dict[tuple, SatRegistro] = {}
    for reg in SatRegistro.objects.filter(empresa=empresa).only("competencia", "sheet", "row"):
        exist_map[(empresa.id, reg.competencia, reg.sheet, reg.row)] = reg

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue

        header_keys = [_slugify_field(h) for h in headers]

        for r, row in enumerate(rows, start=2):
            data, vazio = {}, True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    vazio = False
                data[key] = None if val in (None, "") else str(val)

            if vazio:
                ignorados_vazios += 1
                continue

            desc  = (data.get("descricao") or data.get("descricao_produto") or data.get("descricao_mercadoria"))
            cst   = (data.get("cst_csosn") or data.get("cst") or data.get("csosn"))
            dt_em = _extr_data_emissao_dict(data)
            competencia = _competencia_from_date(dt_em) or comp_param  # pode ser None

            unique_key = (empresa.id, competencia, sheet_name, r)

            if unique_key in exist_map:
                reg = exist_map[unique_key]
                reg.data = data
                reg.descricao = desc
                reg.ncm = data.get("ncm")
                reg.cfop = data.get("cfop")
                reg.cest = data.get("cest")
                reg.cst_csosn = cst
                reg.data_emissao = dt_em
                reg.competencia  = competencia
                to_update.append(reg)
                atualizados += 1
            else:
                obj = SatRegistro(
                    empresa=empresa,
                    sheet=sheet_name,
                    row=r,
                    data=data,
                    descricao=desc,
                    ncm=data.get("ncm"),
                    cfop=data.get("cfop"),
                    cest=data.get("cest"),
                    cst_csosn=cst,
                    data_emissao=dt_em,
                    competencia=competencia,
                )
                to_create.append(obj)
                criados += 1
                exist_map[unique_key] = obj

            if len(to_create) >= BATCH:
                SatRegistro.objects.bulk_create(to_create, batch_size=BATCH, ignore_conflicts=True)
                to_create.clear()
            if len(to_update) >= BATCH:
                SatRegistro.objects.bulk_update(
                    to_update,
                    fields=["data","descricao","ncm","cfop","cest","cst_csosn","data_emissao","competencia"],
                    batch_size=BATCH,
                )
                to_update.clear()

    if to_create:
        SatRegistro.objects.bulk_create(to_create, batch_size=BATCH, ignore_conflicts=True)
    if to_update:
        SatRegistro.objects.bulk_update(
            to_update,
            fields=["data","descricao","ncm","cfop","cest","cst_csosn","data_emissao","competencia"],
            batch_size=BATCH,
        )

    duracao = round(time.time() - inicio, 2)
    messages.success(
        request,
        f"Importação concluída para '{empresa.nome}'. "
        f"Criados: {criados} • Atualizados: {atualizados} • Ignorados vazios: {ignorados_vazios} • Tempo: {duracao}s."
    )
    return redirect("painel_inicial")


# ============================================================
# Comparação Questor × SAT
# ============================================================

@login_required
def comparar_questor_form(request):
    fim_default = date.today()
    inicio_default = fim_default - timedelta(days=6)
    empresas = Empresa.objects.all().order_by("nome")
    return render(request, "comparar_questor_form.html", {
        "empresas": empresas,
        "inicio_default": inicio_default,
        "fim_default": fim_default,
    })

@login_required
def comparar_questor_resultado(request):
    data = request.session.get("questor_cmp")
    if not data:
        messages.error(request, "Nenhum resultado para exibir. Envie a planilha do Questor primeiro.")
        return redirect("comparar_questor_form")

    empresa = get_object_or_404(Empresa, pk=data["empresa_id"])
    erros = data.get("resultado", [])

    return render(request, "comparar_questor_resultado.html", {
        "empresa": empresa,
        "arquivo_nome": data.get("arquivo_nome"),
        "linhas_lidas": data.get("linhas_lidas", 0),
        "pareadas": data.get("pareadas", 0),
        "nao_encontradas": data.get("nao_encontradas", 0),
        "candidatos_cancelados": data.get("candidatos_cancelados", 0),
        "duracao": data.get("duracao", 0),
        "total_erros": len(erros),
        "erros": erros[:500],
        "sat_linhas_periodo": data.get("sat_linhas_periodo", 0),
        "questor_linhas_periodo": data.get("questor_linhas_periodo", 0),
        "inicio": data.get("inicio") or None,
        "fim": data.get("fim") or None,
        "totais_questor": data.get("totais_questor", {}),
        "totais_sat": data.get("totais_sat", {}),
        "sat_autorizadas_fora": data.get("sat_autorizadas_fora", []),
        "sat_autorizadas_fora_total": data.get("sat_autorizadas_fora_total", 0),
    })


def _row_to_norm_dict(headers, row):
    d = { _slug(h): (row[i] if i < len(row) else None) for i, h in enumerate(headers) }
    return {k: ("" if v is None else str(v).strip()) for k, v in d.items()}

CANDS_VALOR_NOTA = [
    "valor_total_nota","valor_total_notas","valor_total","valortotal","vl_total","valor",
    "valornfe","valor_nfce","valor_total_nfe"
]
CANDS_BC_ICMS = [
    "bc_icms","base_icms","base_de_calculo_icms","basecalc_icms","valor_bc_icms","basecalculo_icms",
]
CANDS_VALOR_ICMS = [
    "valor_icms","vl_icms","vlr_icms","valor_do_icms","vlicms","v_icms",
]

def _extr_decimal_by_keys(d: dict, candidates) -> Decimal:
    k = _first_key(d, candidates)
    return _parse_decimal(d.get(k)) if k else Decimal("0")

def _questor_map_por_documento(ws, dt_ini: date | None = None, dt_fim: date | None = None):
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    cols = { _slug(h): i for i, h in enumerate(headers) }

    cand_num = ["cu_numerodocumento","numerodocumento","numero_documento","n_documento","documento","numero"]
    cand_val = ["valor_contabil","valorcontabil","vl_contabil","vlr_contabil","valor_total","valortotal","vl_total","valor"]
    cand_ser = ["serie","serie_documento","serienfe","serie_nfce"]
    cand_dt  = [
        "dataemissao","data_emissao","dt_emissao","emissao","emissao_data",
        "data","dtemissao",
        "data_entrada_saida","data_entrada","data_saida","dt_entrada","dt_saida",
        "dataentradasaida"
    ]

    k_num = _first_key(cols, cand_num)
    k_val = _first_key(cols, cand_val)
    k_ser = _first_key(cols, cand_ser)
    k_dt  = _first_key(cols, cand_dt)

    if not (k_num and k_val):
        raise ValueError("Planilha do Questor sem colunas NumeroDocumento/Valor Contábil (ou Valor Total).")

    doc_exato, doc_digits = {}, {}
    pair_exato, pair_digits = {}, {}

    lidas_total = 0
    lidas_periodo = 0

    for _r, row in enumerate(rows, start=2):
        lidas_total += 1

        dt_row = None
        if k_dt is not None and cols[k_dt] < len(row):
            dt_row = _parse_date_any(row[cols[k_dt]])
        if dt_ini and dt_fim:
            if dt_row is None or not (dt_ini <= dt_row <= dt_fim):
                continue

        dline = _row_to_norm_dict(headers, row)
        if _is_nfe_questor(dline):
            continue

        lidas_periodo += 1

        raw_num = row[cols[k_num]] if cols[k_num] < len(row) and row[cols[k_num]] is not None else ""
        if str(raw_num).strip() == "":
            continue

        val_q = _parse_decimal(row[cols[k_val]] if cols[k_val] < len(row) else "0")
        serie = ""
        if k_ser is not None and cols[k_ser] < len(row):
            v = row[cols[k_ser]]
            serie = "" if v is None else str(v)

        doc_raw, doc_dig = _norm_doc(str(raw_num))
        key_pair_raw, key_pair_dig = _norm_pair(str(raw_num), serie)

        if serie:
            pair_exato[key_pair_raw]  = pair_exato.get(key_pair_raw,  Decimal("0")) + val_q
            pair_digits[key_pair_dig] = pair_digits.get(key_pair_dig, Decimal("0")) + val_q
        else:
            doc_exato[doc_raw]  = doc_exato.get(doc_raw,  Decimal("0")) + val_q
            doc_digits[doc_dig] = doc_digits.get(doc_dig, Decimal("0")) + val_q

    meta = {
        "linhas_lidas_total": lidas_total,
        "linhas_lidas_periodo": lidas_periodo,
        "col_num": k_num,
        "col_val": k_val,
        "col_serie": k_ser or "",
        "col_data": k_dt or "",
    }
    return (doc_exato, doc_digits, pair_exato, pair_digits), meta


def _questor_totais_periodo(ws, dt_ini: date | None, dt_fim: date | None) -> dict:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    cols = { _slug(h): i for i, h in enumerate(headers) }

    k_num = _first_key(cols, ["cu_numerodocumento","numerodocumento","numero_documento","n_documento","documento","numero"])
    k_ser = _first_key(cols, ["serie","serie_documento","serienfe","serie_nfce"])
    k_dt  = _first_key(cols, [
        "dataemissao","data_emissao","dt_emissao","emissao","emissao_data",
        "data","dtemissao","data_entrada_saida","data_entrada","data_saida","dt_entrada","dt_saida","dataentradasaida"
    ])

    if not k_num:
        return {"valor_total": Decimal("0"), "bc_icms": Decimal("0"), "valor_icms": Decimal("0")}

    por_doc_valor, por_doc_bc, por_doc_icms = {}, {}, {}

    for row in rows:
        if dt_ini and dt_fim:
            if k_dt is None or cols[k_dt] >= len(row):
                continue
            dt_row = _parse_date_any(row[cols[k_dt]])
            if dt_row is None or not (dt_ini <= dt_row <= dt_fim):
                continue

        dline = _row_to_norm_dict(headers, row)
        if _is_nfe_questor(dline):
            continue

        num = row[cols[k_num]] if cols[k_num] < len(row) else ""
        if str(num).strip() == "":
            continue

        serie = ""
        if k_ser is not None and cols[k_ser] < len(row):
            s = row[cols[k_ser]]
            serie = "" if s is None else str(s)

        key_raw, _ = _norm_pair(str(num), serie)

        v_nota = _extr_decimal_by_keys(dline, CANDS_VALOR_NOTA)
        v_bc   = _extr_decimal_by_keys(dline, CANDS_BC_ICMS)
        v_icms = _extr_decimal_by_keys(dline, CANDS_VALOR_ICMS)

        por_doc_valor[key_raw] = por_doc_valor.get(key_raw, Decimal("0")) + v_nota
        por_doc_bc[key_raw]    = por_doc_bc.get(key_raw,    Decimal("0")) + v_bc
        por_doc_icms[key_raw]  = por_doc_icms.get(key_raw,  Decimal("0")) + v_icms

    return {
        "valor_total": sum(por_doc_valor.values(), Decimal("0")),
        "bc_icms":     sum(por_doc_bc.values(),    Decimal("0")),
        "valor_icms":  sum(por_doc_icms.values(),  Decimal("0")),
    }

def _sat_totais_periodo(qs, dt_ini: date | None, dt_fim: date | None) -> dict:
    tot_valor = Decimal("0"); tot_bc = Decimal("0"); tot_icms = Decimal("0")
    vistos = set()
    for reg in qs.iterator():
        d = reg.data or {}
        if dt_ini and dt_fim:
            dt_em = _extr_data_emissao_dict(d)
            if dt_em is None or not (dt_ini <= dt_em <= dt_fim):
                continue
        num_sat = _numero_documento_sat(d)
        if not num_sat:
            continue
        serie_sat = d.get(_first_key(d, ["serie","serienfe","serie_nfce"]), "") or ""
        key_pair_raw, _ = _norm_pair(num_sat, serie_sat)
        if key_pair_raw in vistos:
            continue
        vistos.add(key_pair_raw)
        v_nota = _extr_decimal_by_keys(d, CANDS_VALOR_NOTA)
        v_bc   = _extr_decimal_by_keys(d, CANDS_BC_ICMS)
        v_icms = _extr_decimal_by_keys(d, CANDS_VALOR_ICMS)
        tot_valor += v_nota; tot_bc += v_bc; tot_icms += v_icms
    return {"valor_total": tot_valor, "bc_icms": tot_bc, "valor_icms": tot_icms}


@login_required
@require_POST
def comparar_questor(request):
    empresa_id = request.POST.get("empresa_id")
    arq = request.FILES.get("arquivo")
    dt_ini = _parse_date_any(request.POST.get("inicio"))
    dt_fim = _parse_date_any(request.POST.get("fim"))

    if not empresa_id:
        messages.error(request, "Selecione a empresa.")
        return redirect("comparar_questor_form")
    if not arq or not arq.name.lower().endswith(".xlsx"):
        messages.error(request, "Envie um arquivo .xlsx do Questor.")
        return redirect("comparar_questor_form")
    if (dt_ini and dt_fim) and dt_ini > dt_fim:
        messages.error(request, "Período inválido: data inicial maior que a final.")
        return redirect("comparar_questor_form")

    empresa = get_object_or_404(Empresa, pk=empresa_id)

    try:
        wb_q = load_workbook(filename=arq, data_only=True, read_only=True)
    except Exception as e:
        messages.error(request, f"Não consegui abrir o Excel do Questor: {e}")
        return redirect("comparar_questor_form")

    t0 = time.time()
    ws_q = wb_q[wb_q.sheetnames[0]]

    try:
        (q_doc_raw, q_doc_dig, q_pair_raw, q_pair_dig), meta = _questor_map_por_documento(ws_q, dt_ini, dt_fim)
    except ValueError as err:
        messages.error(request, str(err))
        return redirect("comparar_questor_form")

    ws_q = wb_q[wb_q.sheetnames[0]]
    totais_questor = _questor_totais_periodo(ws_q, dt_ini, dt_fim)

    set_pair_raw = set(q_pair_raw.keys())
    set_pair_dig = set(q_pair_dig.keys())
    set_doc_raw  = set(q_doc_raw.keys())
    set_doc_dig  = set(q_doc_dig.keys())

    if not (set_pair_raw or set_pair_dig or set_doc_raw or set_doc_dig):
        messages.warning(request, "A planilha do Questor não trouxe nenhum Documento legível para pareamento no período.")
        return redirect("comparar_questor_form")

    qs_sat = SatRegistro.objects.filter(empresa=empresa).only("data","sheet","row")
    qs_sat_tot = SatRegistro.objects.filter(empresa=empresa).only("data")
    totais_sat = _sat_totais_periodo(qs_sat_tot, dt_ini, dt_fim)

    divergencias = []
    candidatos = pareadas = nao_encontradas = 0
    sat_considerados = 0
    sat_total_periodo = 0

    # Canceladas pareáveis
    for reg in qs_sat.iterator():
        d = reg.data or {}
        dt_em = _extr_data_emissao_dict(d)
        if dt_ini and dt_fim:
            if dt_em is None or not (dt_ini <= dt_em <= dt_fim):
                continue
        sat_total_periodo += 1
        num_sat = _numero_documento_sat(d)
        if not num_sat:
            continue
        serie_sat = d.get(_first_key(d, ["serie","serienfe","serie_nfce"]), "") or ""
        raw_doc, dig_doc = _norm_doc(num_sat)
        key_pair_raw, key_pair_dig = _norm_pair(num_sat, serie_sat)
        if not ((key_pair_raw and key_pair_raw in set_pair_raw) or
                (key_pair_dig and key_pair_dig in set_pair_dig) or
                (raw_doc and raw_doc in set_doc_raw) or
                (dig_doc and dig_doc in set_doc_dig)):
            continue
        sat_considerados += 1
        status = _extr_status_sat(d)
        if not _is_cancelado(status):
            continue
        candidatos += 1

        val_q = (q_pair_raw.get(key_pair_raw) or
                 q_pair_dig.get(key_pair_dig) or
                 q_doc_raw.get(raw_doc) or
                 q_doc_dig.get(dig_doc))
        if val_q is None:
            nao_encontradas += 1
            continue
        pareadas += 1
        if Decimal(val_q) != 0:
            divergencias.append({
                "linha_excel": "-",
                "documento": num_sat,
                "serie": serie_sat,
                "data_emissao": dt_em.isoformat() if dt_em else "",
                "valor_questor": str(val_q),
                "status_sat": _status_legivel(d),
                "valor_sat": str(_extr_valor_total_sat(d)),
                "id_sat": _extr_id_sat(d),
                "sheet_sat": reg.sheet,
                "row_sat": reg.row,
            })

    # Autorizadas fora do Questor
    autorizadas_fora = []
    vistos_aut = set()
    for reg in qs_sat.iterator():
        d = reg.data or {}
        dt_em = _extr_data_emissao_dict(d)
        if dt_ini and dt_fim:
            if dt_em is None or not (dt_ini <= dt_em <= dt_fim):
                continue
        num_sat = _numero_documento_sat(d)
        if not num_sat:
            continue
        serie_sat = d.get(_first_key(d, ["serie","serienfe","serie_nfce"]), "") or ""
        raw_doc, dig_doc = _norm_doc(num_sat)
        key_pair_raw, key_pair_dig = _norm_pair(num_sat, serie_sat)
        status = _extr_status_sat(d)
        if not _is_autorizado(status):
            continue
        if ((key_pair_raw and key_pair_raw in set_pair_raw) or
            (key_pair_dig and key_pair_dig in set_pair_dig) or
            (raw_doc and raw_doc in set_doc_raw) or
            (dig_doc and dig_doc in set_doc_dig)):
            continue
        uniq = key_pair_raw or raw_doc or dig_doc
        if uniq in vistos_aut:
            continue
        vistos_aut.add(uniq)
        autorizadas_fora.append({
            "documento": num_sat,
            "serie": serie_sat,
            "data_emissao": dt_em.isoformat() if dt_em else "",
            "status_sat": _status_legivel(d),
            "valor_sat": str(_extr_valor_total_sat(d)),
            "id_sat": _extr_id_sat(d),
            "sheet_sat": reg.sheet,
            "row_sat": reg.row,
        })

    duracao = round(time.time() - t0, 2)
    request.session["questor_cmp"] = {
        "empresa_id": empresa.id,
        "arquivo_nome": arq.name,
        "linhas_lidas": meta["linhas_lidas_periodo"],
        "pareadas": pareadas,
        "nao_encontradas": nao_encontradas,
        "candidatos_cancelados": candidatos,
        "duracao": duracao,
        "resultado": divergencias[:20000],
        "meta": meta,
        "sat_linhas_periodo": sat_total_periodo,
        "questor_linhas_periodo": meta["linhas_lidas_periodo"],
        "sat_considerados": sat_considerados,
        "inicio": dt_ini.isoformat() if dt_ini else "",
        "fim": dt_fim.isoformat() if dt_fim else "",
        "totais_questor": {k: str(v) for k, v in (totais_questor or {}).items()},
        "totais_sat": {k: str(v) for k, v in (totais_sat or {}).items()},
        "sat_autorizadas_fora": autorizadas_fora[:20000],
        "sat_autorizadas_fora_total": len(autorizadas_fora),
    }

    messages.success(
        request,
        f"Período: {dt_ini or '—'} a {dt_fim or '—'} • "
        f"SAT no período: {sat_total_periodo} • Presentes no Questor: {sat_considerados} • "
        f"Canceladas: {candidatos} • Pareadas: {pareadas} • Não encontradas: {nao_encontradas} • "
        f"Divergências (valor>0): {len(divergencias)} • "
        f"SAT autorizadas fora do Questor: {len(autorizadas_fora)}."
    )
    return redirect("comparar_questor_resultado")


# ------------------ Exportação XLSX ------------------

@login_required
def comparar_questor_csv(request):
    data = request.session.get("questor_cmp")
    if not data:
        messages.error(request, "Nenhum resultado de comparação encontrado na sessão.")
        return redirect("comparar_questor_form")

    empresa = get_object_or_404(Empresa, pk=data["empresa_id"])
    rows = data.get("resultado", []) or []
    sat_aut = data.get("sat_autorizadas_fora", []) or []

    def _autofit(ws, min_w=9, max_w=80):
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, (datetime, date)):
                    s = v.strftime("%d/%m/%Y")
                else:
                    s = str(v)
                if len(s) > 200:
                    s = s[:200]
                max_len = max(max_len, len(s))
                cell.alignment = Alignment(wrap_text=False, vertical="center")
            ws.column_dimensions[letter].width = max(min_w, min(int(max_len * 1.15), max_w))

    def _to_dec(x):
        try:
            s = str(x).replace(".", "").replace(",", ".")
            return float(Decimal(s))
        except Exception:
            try:
                return float(Decimal(str(x)))
            except Exception:
                return 0.0

    def _to_date(x):
        return _parse_date_any(x)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    status_cancel_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    status_ok_fill     = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")

    wb = Workbook()

    # Aba 1: Divergências (com resumo)
    ws = wb.active
    ws.title = "Divergências"

    ws.append(["RESUMO DA ANÁLISE"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Empresa", empresa.nome])
    ws.append(["CNPJ", empresa.cnpj])
    ws.append(["Arquivo Questor", data.get("arquivo_nome")])
    ws.append(["Período Início", data.get("inicio") or "—"])
    ws.append(["Período Fim", data.get("fim") or "—"])
    ws.append([])

    ws.append(["ESTATÍSTICAS"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    ws.append(["Total SAT no período", data.get("sat_linhas_periodo")])
    ws.append(["Total Questor no período", data.get("questor_linhas_periodo")])
    ws.append(["SAT presentes no Questor", data.get("sat_considerados")])
    ws.append(["Canceladas no SAT", data.get("candidatos_cancelados")])
    ws.append(["Pareadas", data.get("pareadas")])
    ws.append(["Não encontradas", data.get("nao_encontradas")])
    ws.append(["❌ Divergências (valor>0)", len(rows)])
    ws.append(["SAT autorizadas fora do Questor", data.get("sat_autorizadas_fora_total", 0)])
    ws.append(["Tempo processamento (s)", data.get("duracao")])
    ws.append([])

    tq = data.get("totais_questor", {}) or {}
    ts = data.get("totais_sat", {}) or {}
    ws.append(["TOTAIS NO PERÍODO"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)

    header_row = ws.max_row + 1
    ws.append(["", "Valor total notas", "BC ICMS", "Valor ICMS"])
    for col in range(1, 5):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align

    ws.append(["Questor", _to_dec(tq.get("valor_total","0")), _to_dec(tq.get("bc_icms","0")), _to_dec(tq.get("valor_icms","0"))])
    ws.append(["SAT",     _to_dec(ts.get("valor_total","0")), _to_dec(ts.get("bc_icms","0")), _to_dec(ts.get("valor_icms","0"))])
    for r in (header_row+1, header_row+2):
        for c in ("B","C","D"):
            ws[f"{c}{r}"].number_format = "#,##0.00"

    ws.append([])

    ws.append(["DIVERGÊNCIAS - NFC-e CANCELADAS COM VALOR CONTÁBIL"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)

    header_row = ws.max_row + 1
    ws.append([
        "Número Documento", "Série", "Data Emissão", "Valor Contábil Questor",
        "Status SAT", "Valor Total SAT", "ID SAT", "Aba SAT", "Linha SAT"
    ])
    for col in range(1, 10):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for e in rows:
        dt = _to_date(e.get("data_emissao",""))
        ws.append([
            e.get("documento",""), e.get("serie",""), dt,
            _to_dec(e.get("valor_questor","0")),
            e.get("status_sat",""), _to_dec(e.get("valor_sat","0")),
            e.get("id_sat",""), e.get("sheet_sat",""), e.get("row_sat","")
        ])
        rix = ws.max_row
        st = str(ws.cell(row=rix, column=5).value or "").upper()
        if st == "CANCELADA":
            ws.cell(row=rix, column=5).fill = status_cancel_fill
        elif st == "AUTORIZADA":
            ws.cell(row=rix, column=5).fill = status_ok_fill

    # formatos e layout
    for cell in ws.iter_cols(min_col=3, max_col=3, min_row=header_row+1, max_row=ws.max_row):
        for c in cell: c.number_format = "dd/mm/yyyy"
    for cell in ws.iter_cols(min_col=4, max_col=4, min_row=header_row+1, max_row=ws.max_row):
        for c in cell: c.number_format = "#,##0.00"
    for cell in ws.iter_cols(min_col=6, max_col=6, min_row=header_row+1, max_row=ws.max_row):
        for c in cell: c.number_format = "#,##0.00"

    ws.auto_filter.ref = f"A{header_row}:I{ws.max_row}"
    ws.freeze_panes = f"A{header_row+1}"
    # auto fit e coluna ID maior
    def _autofit(ws, min_w=9, max_w=80):
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, (datetime, date)):
                    s = v.strftime("%d/%m/%Y")
                else:
                    s = str(v)
                if len(s) > 200:
                    s = s[:200]
                max_len = max(max_len, len(s))
                cell.alignment = Alignment(wrap_text=False, vertical="center")
            ws.column_dimensions[letter].width = max(min_w, min(int(max_len * 1.15), max_w))
    _autofit(ws)
    ws.column_dimensions['G'].width = max(ws.column_dimensions['G'].width or 0, 52)

    # Aba 2: SAT autorizadas fora
    ws2 = wb.create_sheet("SAT autorizadas fora")
    header_row2 = 1
    ws2.append(["Número Documento","Série","Data Emissão","Status SAT",
                "Valor Total SAT","ID SAT","Aba SAT","Linha SAT"])
    for col in range(1, 9):
        cell = ws2.cell(row=header_row2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for e in sat_aut:
        ws2.append([
            e.get("documento",""),
            e.get("serie",""),
            _to_date(e.get("data_emissao","")),
            e.get("status_sat",""),
            _to_dec(e.get("valor_sat","0")),
            e.get("id_sat",""),
            e.get("sheet_sat",""),
            e.get("row_sat",""),
        ])
        rix = ws2.max_row
        st = str(ws2.cell(row=rix, column=4).value or "").upper()
        if st == "AUTORIZADA":
            ws2.cell(row=rix, column=4).fill = status_ok_fill

    for cell in ws2.iter_cols(min_col=3, max_col=3, min_row=2, max_row=ws2.max_row):
        for c in cell: c.number_format = "dd/mm/yyyy"
    for cell in ws2.iter_cols(min_col=5, max_col=5, min_row=2, max_row=ws2.max_row):
        for c in cell: c.number_format = "#,##0.00"

    ws2.auto_filter.ref = f"A1:H{ws2.max_row}"
    ws2.freeze_panes = "A2"
    _autofit(ws2)
    ws2.column_dimensions['F'].width = max(ws2.column_dimensions['F'].width or 0, 52)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"divergencias_questor_{empresa.cnpj}.xlsx".replace("/", "_")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp
